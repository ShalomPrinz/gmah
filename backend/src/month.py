from enum import Enum
from glob import glob
from os import path
from openpyxl import load_workbook

from src.data import key_prop, report_properties, date_prop, status_prop
from src.errors import FileAlreadyExists
from src.excel import Excel
from src.families import load_families_file, search_families
from src.managers import find_manager, load_managers_file
from src.styles import report_cell_style, report_received_style, report_not_received_style, report_received_name, report_not_received_name, style_name

class ReportSearchBy(Enum):
    NAME = 'name'
    MANAGER = 'manager'
    DRIVER = 'driver'
    RECEIVE = 'recieve'

    @classmethod
    def get_search_columns(cls, search_by):
        if search_by is None:
            return []
        
        search_by = getattr(ReportSearchBy, search_by.upper(), ReportSearchBy.NAME)
        match search_by:
            case ReportSearchBy.NAME:
                return [0]
            case ReportSearchBy.MANAGER:
                return [1]
            case ReportSearchBy.DRIVER:
                return [2]
            case ReportSearchBy.RECEIVE:
                return [4]
            case _:
                return [0]

report_style_map = {
    style_name: None,
    report_received_name: True,
    report_not_received_name: False
}

month_reports_folder = "דוחות קבלה"
month_reports_path = f"{month_reports_folder}/"
month_reports_template = f"{month_reports_path}template.xlsx"

month_report_prefix = "דוח קבלה "
month_report_suffix = ".xlsx"
month_reports_pattern = f"{month_reports_path}{month_report_prefix}*{month_report_suffix}"

default_driver = ""
default_manager = ""

def get_report_path(report_name):
    '''
    Returns report path from project parent directory.
    '''
    return f'{month_reports_path}{month_report_prefix}{report_name}{month_report_suffix}'

def is_report_name_exists(report_name):
    '''
    Validates if a report already exists in generated reports.
    '''
    filepath = get_report_path(report_name)
    return not path.isfile(filepath)

def load_report_file(report_name):
    '''
    Connects to a report file.

    Returns a tuple: (error, file)
        - If connection has failed, file will be None
        - If connection has succeed, error will be None
    '''
    path = get_report_path(report_name)
    try:
        report = Excel(
            filename=path,
            row_properties=report_properties,
            search_enum=ReportSearchBy,
            required_style=report_cell_style
        )
        report.add_named_style(report_received_style)
        report.add_named_style(report_not_received_style)
        return (None, report)
    except Exception as e:
        return (e, None)

def load_template(template_path, sheet_title):
    '''
    Loads a template, modifies its sheet title and returns the workbook.
    '''
    workbook = load_workbook(template_path)
    sheet = workbook[workbook.sheetnames[0]]
    sheet.title = sheet_title
    return workbook

def create_from_template(name):
    '''
    Creates empty month report.
    '''
    sheet_title = f'{month_report_prefix}{name}'
    workbook = load_template(month_reports_template, sheet_title)
    filepath = get_report_path(name)
    workbook.save(filepath)

def to_excel_row(family, managers_file):
    '''
    Cast family data to report excel row format in the right order.
    If family is missing key_prop, detailed exception is raised.
    If family is missing driver, default_driver will be their driver.
    '''
    family_key = family.get(key_prop, None)
    if family_key is None:
        raise Exception(f"שגיאה ביצירת דוח קבלה חודשי: למשפחה {family} אין שם")
    
    driver = family.get("נהג", default_driver)
    manager = find_manager(managers_file, driver) or default_manager

    return [family_key, manager, driver, None, None]

def generate_month_report(name, override_name=False):
    '''
    Generates new month report with the given name, based on current families and managers files.
    Allows override of exist report by setting override_name to True.
    '''
    if not override_name and not is_report_name_exists(name):
        return FileAlreadyExists(f"דוח קבלה בשם {name} קיים כבר")
    
    error, families_file = load_families_file()
    if error is not None:
        return error
    
    error, managers_file = load_managers_file()
    if error is not None:
        return error
    
    create_from_template(name)
    error, report_file = load_report_file(name)
    if error is not None:
        return error
    
    family_to_excel_row = lambda family: to_excel_row(family, managers_file)
    report_file.append_rows(search_families(families_file), family_to_excel_row)

def get_no_manager_drivers():
    '''
    Returns all the drivers from families file that have no corresponding manager in managers file.
    '''
    error, families_file = load_families_file()
    if error is not None:
        return error, None
    
    error, managers_file = load_managers_file()
    if error is not None:
        return error, None
    
    no_manager_drivers = []

    for family in search_families(families_file):
        family_key = family.get(key_prop, None)
        driver = family.get("נהג", None)
        if family_key is None or driver is None:
            continue

        manager = find_manager(managers_file, driver)
        if manager is None:
            for d in no_manager_drivers:
                if d["name"] == driver:
                    d["count"] += 1
                    break
            else:
                no_manager_drivers.append({ "name": driver, "count": 1 })

    return None, no_manager_drivers

def get_no_driver_families():
    '''
    Returns all the families from families file that doesn't have a driver.
    '''
    error, families_file = load_families_file()
    if error is not None:
        return error, None
    
    no_driver_families = 0

    for family in search_families(families_file):
        family_key = family.get(key_prop, None)
        if family_key is None:
            continue
        
        driver = family.get("נהג", None)
        if driver is None:
            no_driver_families += 1

    return None, no_driver_families

def get_reports_list():
    '''
    Returns a list with all the generated monthly reports.
    '''
    start_index = len(month_report_prefix)
    end_index = len(month_report_suffix)
    return [path.basename(file)[start_index:-end_index] for file in glob(month_reports_pattern)]

def search_report(report_file: Excel, query='', search_by=''):
    '''
    Returns list of families from the given report_file,
    who their value of the search_by cell matches the given query
    '''
    query = '' if query is None else query
    search_by = '' if search_by is None else search_by

    return report_file.style_search(query, search_by, search_style='receive', style_map=report_style_map)

def search_report_column(report_file: Excel, query='', search_by=''):
    '''
    Returns list of families value of the given search_by column from the report_file.
    '''
    query = '' if query is None else query
    search_by = '' if search_by is None else search_by

    return report_file.column_search(query, search_by)

def get_receipt_status(report_file: Excel, family_name):
    '''
    Returns receipt status of family_name in report_file.
    If family_name not found, returns default receipt status.
    '''    
    default_date = ""
    default_status = False

    report = report_file.style_search(family_name, 'name', search_style='receive', style_map=report_style_map)
    family_report_data = report[0] if report else {}

    return {
        "date": family_report_data.get(date_prop, default_date),
        "status": family_report_data.get(status_prop, default_status)
    }

def update_receipt_status(report_file: Excel, family_name, receipt):
    '''
    Updates receipt status of family_name in report_file to be the given receipt.
    '''
    try:
        index = report_file.get_row_index(family_name)
    except Exception as e:
        return e
    
    if (date := receipt.get("date", None)) is not None:
        report_file.replace_cell(index, {
            "key": date_prop,
            "value": date
        })

    if (status := receipt.get("status", None)) is not None:
        style = report_received_name if status else report_not_received_name
        report_file.replace_cell(index, {
            "key": status_prop,
            "style": style
        })
