from enum import Enum
from openpyxl import load_workbook
from re import match

from src.data import report_properties, key_prop, street_prop, driver_prop, date_prop, status_prop, date_pattern, default_date, default_status
from src.errors import FamilyNotFoundError
from src.excel import Excel
from src.families import search_families
from src.json import Json
from src.managers import find_manager
from src.results import receipt_update_results
from src.styles import report_cell_style, report_received_style, report_not_received_style, report_received_name, report_not_received_name, style_name

class ReportSearchBy(Enum):
    NAME = 'name'
    MANAGER = 'manager'
    DRIVER = 'driver'
    RECEIVE = 'recieve'

    @classmethod
    def get_search_columns(cls, search_by):
        if search_by is None:
            return []

        search_by = getattr(
            ReportSearchBy,
            search_by.upper(),
            ReportSearchBy.NAME)
        match search_by:
            case ReportSearchBy.NAME:
                return [0]
            case ReportSearchBy.MANAGER:
                return [1]
            case ReportSearchBy.DRIVER:
                return [2]
            case ReportSearchBy.RECEIVE:
                return [4]
            case _:
                return [0]

report_style_map = {
    style_name: None,
    report_received_name: True,
    report_not_received_name: False
}

default_driver = ""
default_manager = ""
default_receipt = {
    "date": default_date,
    "status": default_status
}

def load_report_file(path):
    '''
    Connects to a report file.

    Returns a tuple: (error, file)
        - If connection has failed, file will be None
        - If connection has succeed, error will be None
    '''
    try:
        report = Excel(
            filename=path,
            row_properties=report_properties,
            search_enum=ReportSearchBy,
            required_style=report_cell_style
        )
        report.add_named_style(report_received_style)
        report.add_named_style(report_not_received_style)
        return (None, report)
    except Exception as e:
        return (e, None)

# Report preparation and validation

def get_no_manager_drivers(families_file: Excel, managers_file: Json):
    '''
    Returns all the drivers from families file that have no corresponding manager in managers file.
    '''
    no_manager_drivers = []

    for family in search_families(families_file):
        family_key = family.get(key_prop, None)
        driver = family.get("נהג", None)
        if family_key is None or driver is None:
            continue

        manager = find_manager(managers_file, driver)
        if manager is None:
            for d in no_manager_drivers:
                if d["name"] == driver:
                    d["count"] += 1
                    break
            else:
                no_manager_drivers.append({ "name": driver, "count": 1 })

    return None, no_manager_drivers

def get_no_driver_families(families_file: Excel):
    '''
    Returns all the families from families file that doesn't have a driver.
    '''
    no_driver_families = 0

    for family in search_families(families_file):
        family_key = family.get(key_prop, None)
        if family_key is None:
            continue

        driver = family.get("נהג", None)
        if driver is None:
            no_driver_families += 1

    return None, no_driver_families

# Report generation

def create_empty_report(template_path, sheet_title, filepath):
    '''
    Loads a template, modifies its sheet title and saves the workbook.
    When this function returns, given filepath is a blank month report.
    '''
    workbook = load_workbook(template_path)
    sheet = workbook[workbook.sheetnames[0]]
    sheet.title = sheet_title
    workbook.save(filepath)

def append_report(report_file: Excel, families, managers_file):
    '''
    Appends all given families to the given report.
    '''
    def family_to_excel_row(family): return to_excel_row(family, managers_file)
    report_file.append_rows(families, family_to_excel_row)

def report_late_append(report_file: Excel, families):
    '''
    Appends all given families to the given report.
    Families won't have driver or manager inside report.
    '''
    def family_to_excel_row(family):
        family_key = family.get(key_prop, None)
        if family_key is None:
            raise Exception(f"שגיאה בעדכון דוח קבלה חודשי: למשפחה {family} אין שם")
        return [family_key, None, None, None, None]
    report_file.append_rows(families, family_to_excel_row)

def to_excel_row(family, managers_file):
    '''
    Cast family data to report excel row format in the right order.
    If family is missing key_prop, detailed exception is raised.
    If family is missing driver, default_driver will be their driver.
    '''
    family_key = family.get(key_prop, None)
    if family_key is None:
        raise Exception(f"שגיאה ביצירת דוח קבלה חודשי: למשפחה {family} אין שם")

    driver = family.get("נהג", default_driver)
    manager = find_manager(managers_file, driver) or default_manager

    return [family_key, manager, driver, None, None]

# Report data tracking

def search_report(report_file: Excel, query='', search_by=''):
    '''
    Returns list of families from the given report_file,
    who their value of the search_by cell matches the given query
    '''
    query = '' if query is None else query
    search_by = '' if search_by is None else search_by

    return report_file.style_search(
        query,
        search_by,
        search_style='receive',
        style_map=report_style_map)

def search_report_column(report_file: Excel, query='', search_by=''):
    '''
    Returns list of families value of the given search_by column from the report_file.
    '''
    query = '' if query is None else query
    search_by = '' if search_by is None else search_by

    return report_file.column_search(query, search_by)

def get_report_completion_families(report_file: Excel, families_file: Excel):
    '''
    Returns families who didn't receive a package in the given report_file,
    and only families to whom a package was sent and didn't get it.
    Each completion family has a name, driver and street.
    '''
    families = search_families(families_file)

    def status_filter(family):
        not_received = family[status_prop] == report_style_map[report_not_received_name]
        no_driver_no_status = False if family[driver_prop] else family[status_prop] == report_style_map[style_name]
        return not_received or no_driver_no_status
    filtered_families = filter(status_filter, search_report(report_file))

    def to_completion_family(family):
        if not family[key_prop]:
            return {}

        family_data = next(
            (f for f in families if f[key_prop] == family[key_prop]), None)
        return {} if family_data is None else {
            key_prop: family[key_prop],
            street_prop: family_data[street_prop],
            driver_prop: family[driver_prop]
        }
    completion_families = map(to_completion_family, filtered_families)

    def not_empty_filter(completion_family):
        return len(completion_family.keys()) > 0

    return list(filter(not_empty_filter, completion_families))

def to_receipt_status(family):
    '''
    Cast a family to receipt status dictionary.
    Should be used for getting data only.
    '''
    return {
        "date": family.get(date_prop, default_date),
        "status": family.get(status_prop, default_status)
    }

def get_driver_receipt_status(report_file: Excel, driver_name):
    '''
    Returns receipt statuses of all families in report_file who their
    driver is driver_name. If driver_name not found, returns empty families list.
    '''
    if not driver_name:
        return []

    result = report_file.style_search(
        driver_name,
        'driver',
        search_style='receive',
        style_map=report_style_map,
        exact=True)

    if not result:
        return []

    def family_map(family):
        if (name := family.get(key_prop, None)) is None:
            return {}

        return {
            **to_receipt_status(family),
            "name": name
        }

    return list(map(family_map, result))

def get_family_receipt_status(report_file: Excel, family_name):
    '''
    Returns receipt status of family_name in report_file.
    If family_name not found, returns default receipt status.
    '''
    if not family_name:
        return default_receipt

    result = report_file.style_search(
        family_name,
        'name',
        search_style='receive',
        style_map=report_style_map)

    return to_receipt_status(result[0]) if result else default_receipt

def validate_date_format(date):
    return True if date and isinstance(
        date, str) and match(
        date_pattern, date) else False

def update_driver_receipt_status(report_file: Excel, status):
    '''
    Updates receipt status of all families in the given status list,
    only if there is a date property in the family's receipt.
    Returns proper Result object.
    '''
    errors = 0
    for family in status:
        name = family.get("name")
        result = update_family_receipt_status(report_file, name, family)
        if result.status != 200:
            errors += 1

    result = "UPDATE_FAILED" if errors == len(status) else (
        "DRIVER_UPDATED" if errors == 0 else "PARTIAL_UPDATE")
    return receipt_update_results[result]

def update_family_receipt_status(report_file: Excel, family_name, receipt):
    '''
    Updates receipt status of family_name in report_file to be the given receipt,
    only if there is a date property in the receipt.
    Returns proper Result object.
    '''
    if not family_name:
        return FamilyNotFoundError(
            "אי אפשר לעדכן סטטוס קבלה של משפחה ללא שם").result

    try:
        index = report_file.get_row_index(family_name)
    except FamilyNotFoundError as e:
        return e.result

    date = receipt.get("date")
    if not date:
        return receipt_update_results["MISSING_DATE"]

    if not validate_date_format(date):
        return receipt_update_results["DATE_MALFORMED"]

    report_file.replace_cell(index, {
        "key": date_prop,
        "value": date
    })

    if (status := receipt.get("status", None)) is not None:
        style = report_received_name if status else report_not_received_name
        report_file.replace_cell(index, {
            "key": status_prop,
            "style": style
        })

    return receipt_update_results["RECEIPT_UPDATED"]
