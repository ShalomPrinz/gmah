from openpyxl import load_workbook
from os import path

from src.search import SearchRequest, search, StyleSearchRequest, style_search, ColumnSearchRequest, search_column, FindRequest, find
from src.errors import FileResourcesMissingError, FamilyNotFoundError
from src.util import letter_by_index
from src.styles import NamedStyle

class Excel:
    def __init__(self, filename: str, row_properties, search_enum,
                 required_style: NamedStyle, table_name: str = ""):
        if not path.exists(filename):
            raise FileNotFoundError(f'הקובץ {filename} לא נמצא')

        self.filename = filename
        self.workbook = load_workbook(filename)
        self.worksheet = self.workbook[self.workbook.sheetnames[0]]

        self.table_name = table_name
        self.cell_style = required_style.name

        self.row_properties = row_properties
        self.search_enum = search_enum
        self.last_column = letter_by_index(len(row_properties))
        self.first_content_row = 2 # First row is 1, and contains titles

        self.add_named_style(required_style)

        if self.table_name and self.table_name not in self.worksheet.tables:
            raise FileResourcesMissingError(
                f"על הקובץ {filename} להכיל טבלה בשם '{self.table_name}'")

    def save(self):
        self.workbook.save(self.filename)

    def get_rows_num(self):
        return self.worksheet.max_row

    def get_headers(self):
        return [cell.value for cell in next(self.worksheet.rows)]

    def get_rows_iter(self):
        return self.worksheet.iter_rows(min_row=self.first_content_row)

    def get_row_index(self, row_key):
        request = FindRequest(
            rows_iter=self.get_rows_iter(),
            query=row_key,
            search_enum=self.search_enum
        )

        result = find(request)
        if result != -1:
            return result + self.first_content_row
        else:
            raise FamilyNotFoundError(f"המשפחה {row_key} לא נמצאת")

    def search(self, query, search_by='', empty=False, exact=False):
        request = SearchRequest(
            rows_iter=self.get_rows_iter(),
            headers=self.get_headers(),
            query=query,
            search_by=search_by,
            search_enum=self.search_enum,
            empty_search=empty,
            exact=exact
        )

        return search(request)

    def style_search(
            self,
            query,
            search_by='',
            search_style=None,
            style_map={}):
        request = StyleSearchRequest(
            headers=self.get_headers(),
            query=query,
            rows_iter=self.get_rows_iter(),
            search_by=search_by,
            search_enum=self.search_enum,
            search_style=search_style,
            style_map=style_map
        )

        return style_search(request)

    def column_search(self, query, search_by=''):
        request = ColumnSearchRequest(
            query=query,
            rows_iter=self.get_rows_iter(),
            search_by=search_by,
            search_enum=self.search_enum
        )

        return search_column(request)

    def append_rows(self, families, to_excel_row):
        for family in families:
            new_row = self.get_rows_num() + 1
            self.worksheet.insert_rows(idx=new_row, amount=1)

            row_data = to_excel_row(family)
            for column, value in enumerate(row_data, 1):
                cell = self.worksheet.cell(row=new_row, column=column)
                cell.value = value
                cell.style = self.cell_style

            if self.table_name:
                self.worksheet.tables[self.table_name].ref = f'A1:{self.last_column}{new_row}'

        self.save()

    def remove_row(self, row_index):
        self.worksheet.delete_rows(row_index)
        self.save()

    def replace_row(self, row_index, row_data):
        for key, value in row_data.items():
            if key not in self.row_properties:
                continue

            col_index = self.row_properties.index(key) + 1
            cell = self.worksheet.cell(row=row_index, column=col_index)
            cell.value = value

        self.save()

    def replace_cell(self, row_index, cell_data):
        '''
        Replaces a single cell with the given cell_data.
        Expects cell_data to be a dict with a "key" attr, in addition to a "style" or "value"
        attributes. If both present, "value" will be ignored.
        '''
        if cell_data["key"] not in self.row_properties:
            return

        col_index = self.row_properties.index(cell_data["key"]) + 1
        cell = self.worksheet.cell(row=row_index, column=col_index)

        if "style" in cell_data:
            cell.style = cell_data["style"]
        elif "value" in cell_data:
            cell.value = cell_data["value"]

        self.save()

    def add_named_style(self, named_style):
        if named_style.name not in self.workbook.named_styles:
            self.workbook.add_named_style(named_style)
            self.save()
