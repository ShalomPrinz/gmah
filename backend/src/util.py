from collections import defaultdict
from openpyxl import load_workbook
from os import makedirs, path, umask
from uuid import uuid4

from src.data import driver_prop
from src.results import driver_update_results

def without_hyphen(string: str):
    return string.replace('-', '')

def insert_hyphen(string: str, index: int):
    return string[:index] + '-' + string[index:]

def letter_by_index(index: int):
    '''
    Index starts at 1. To get the letter A, call the function with index = 1.
    '''
    return chr(ord('A') + index - 1)

def generate_random_id():
    return str(uuid4())

def create_folders_path(folders):
    '''
    If given folders path doesn't exist, creates all folders in path, full permission granted.
    '''
    if path.exists(folders):
        return

    try:
        original_umask = umask(0)
        makedirs(folders, 0o777)
    finally:
        umask(original_umask)

def unique_list(lst):
    return list(dict.fromkeys(lst))

DRIVER_NAME_MIN_LENGTH = 2

def validate_driver_name(driver_name):
    '''
    Validates driver_name is valid and not too short.
    If result is None, validation succeed.
    '''
    if not driver_name or not isinstance(driver_name, str):
        return driver_update_results["MISSING_DRIVER"]

    if len(driver_name) < DRIVER_NAME_MIN_LENGTH:
        return driver_update_results["TOO_SHORT_DRIVER"]
    
    return None

def duplicate_excel_template(template_path, sheet_title, filepath):
    '''
    Loads a template, modifies its sheet title and saves the workbook in filepath.
    '''
    workbook = load_workbook(template_path)
    sheet = workbook[workbook.sheetnames[0]]
    sheet.title = sheet_title
    workbook.save(filepath)

empty_values = [None, ""]

def get_driver_status(driver_name, managers):
    '''
    If given driver has a manager, returns his name.
    If given driver should be print-ignored, returns "ignore".
    If no manager found, returns None.
    '''
    for manager in managers:
        if manager.get('print', None) == "ignore":
            continue
        for driver in manager['drivers']:
            if driver_name == driver['name']:
                if driver.get('print', None) == "ignore":
                    return "ignore"
                return manager['name']
    return None

def sort_families_by_drivers(managers, families):
    '''
    Returns a tuple which contains three lists, generated out of given managers and families.
    This tuple is the result of the sorting phase, before the actual pdf build.
    '''
    managers_families = defaultdict(lambda: defaultdict(list))
    non_managers_families = defaultdict(list)
    driverless_families = []

    for family in families:
        driver_name = family.get(driver_prop, None)
        if driver_name in empty_values:
            driverless_families.append(family)

        driver_status = get_driver_status(driver_name, managers)
        if driver_status == "ignore":
            continue
        elif driver_status is None:
            non_managers_families[driver_name].append(family)
        else:
            managers_families[driver_status][driver_name].append(family)
    
    return managers_families, non_managers_families, driverless_families

def get_all_pages(managers, families):
    '''
    Returns pages list generated out of given managers and families.
    '''
    managers_families, non_managers_families, driverless_families = sort_families_by_drivers(managers, families)

    pages = []
    def append_manager_page(manager_name):
        pages.append({ "title": manager_name })
    def append_driver_page(title, families):
        if title in empty_values or len(families) <= 0:
            return
        pages.append({ "title": title, "content": families })

    # managers_families
    for manager, drivers in managers_families.items():
        append_manager_page(manager)
        for driver, families in drivers.items():
            append_driver_page(driver, families)
    
    # non_managers_families
    for driver, families in non_managers_families.items():
        append_driver_page(driver, families)

    # driverless_families
    append_driver_page("משפחות ללא נהג", driverless_families)
    return pages
