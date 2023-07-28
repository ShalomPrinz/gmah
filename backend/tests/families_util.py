from openpyxl import load_workbook

from src.data import family_properties, families_filename, families_history_filename
from src.families import load_families_file, load_families_history_file

from tests.tests_util import restore_file, store_file

temp_families_filename = 'temp_families.xlsx'
temp_families_history_filename = 'temp_families_history.xlsx'

default_family_properties = {
    "שם מלא": "פלוני פלוני",
    "רחוב": "שפרינצק",
    "בניין": 10,
    "דירה": 2,
    "קומה": 1,
    "מס' בית": "012-3456789",
    "מס' פלאפון": "987-6543210",
    "נהג": "שלמה שלומי",
    "ממליץ": "רווחה",
    "הערות": "",
}

defauly_history_properties = {k: v for i, (k, v) in enumerate(default_family_properties.items()) if i < len(default_family_properties) - 3}
defauly_history_properties.update({
    "ממליץ": "רווחה",
    "תאריך יציאה": "",
    "סיבה": ""
})

class Family:
    def __init__(self, family):
        self.excel_row = [family.get(key, default_family_properties.get(key, None)) for key in family_properties]

class HistoryFamily:
    def __init__(self, family):
        self.excel_row = [family.get(key, defauly_history_properties.get(key, None)) for key in defauly_history_properties]

def write_excel_data(families, filename):
    workbook = load_workbook(filename)
    worksheet = workbook[workbook.sheetnames[0]]
    worksheet.delete_rows(2, worksheet.max_row - 1)
    
    for family in families:
        worksheet.append(family.excel_row)
    workbook.save(filename)

def write_families(families):
    write_excel_data(families, families_filename)

def empty_families():
    write_excel_data([], families_filename)

def write_history_families(families):
    write_excel_data(families, families_history_filename)

def empty_families_history():
    write_excel_data([], families_history_filename)

def load_families():
    error, families_file = load_families_file()
    if error is not None:
        raise Exception("Couldn't load families file", error)
    else:
        return families_file

def load_families_history():
    error, history_file = load_families_history_file()
    if error is not None:
        raise Exception("Couldn't load families history file", error)
    else:
        return history_file

def load_both_families_files():
    return load_families(), load_families_history()

def setUpFamilies():
    store_file(families_filename, temp_families_filename)
    store_file(families_history_filename, temp_families_history_filename)

def tearDownFamilies():
    restore_file(families_filename, temp_families_filename)
    restore_file(families_history_filename, temp_families_history_filename)
