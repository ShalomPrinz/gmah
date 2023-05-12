import openpyxl
import unittest
from shutil import copy
from os import remove

from src.data import family_properties, families_filename
from src.families import get_count, search_families, add_family, add_families, open_families_file
from src.results import add_results, add_many_results, add_many_error

default_family_properties = {
    "שם מלא": "פלוני פלוני",
    "רחוב": "שפרינצק",
    "בניין": 10,
    "דירה": 2,
    "קומה": 1,
    "מס' בית": "012-3456789",
    "מס' פלאפון": "987-6543210",
    "נהג": "שלמה שלומי",
    "נהג במקור": "נחום נחום",
    "ממליץ": "רווחה",
    "הערות": "",
}

class Family:
    def __init__(self, family):
        self.excel_row = [family.get(key, default_family_properties.get(key, None)) for key in family_properties]

def write_families_file(families):
    workbook = openpyxl.load_workbook(families_filename)
    worksheet = workbook[workbook.sheetnames[0]]
    worksheet.delete_rows(2, worksheet.max_row - 1)

    for family in families:
        worksheet.append(family.excel_row)
    workbook.save(families_filename)

def load_families_file():
    error, families_file = open_families_file()
    if error is not None:
        raise Exception("Couldn't load families file", error)
    else:
        return families_file

class TestSearch(unittest.TestCase):
    def setUpClass():
        copy(families_filename, 'temp.xlsx')
    
    def tearDownClass():
        copy('temp.xlsx', families_filename)
        remove('temp.xlsx')

    def test_search_by_name(self):
        families = [Family({"שם מלא": "פרינץ"}), Family({"שם מלא": "כהנא"}), Family({"שם מלא": "נתאי"})]
        
        test_cases = [
            ("None", None, 3),
            ("Empty String", "", 3),
            ("Two Matches", "י", 2),
            ("First Letter", "פ", 1),
            ("Second Letter", "ה", 1),
            ("Not Found", "ברוזוביץ", 0)
        ]

        for title, query, expected_len in test_cases:
            with self.subTest(title=title):
                write_families_file(families=families)
                families_file = load_families_file()
                search_result = search_families(families_file, query)
                self.assertEqual(expected_len, len(search_result))
    
    def test_search_by_street(self):
        families = [Family({"שם מלא": "פרינץ", "רחוב": "הבנים"}),
            Family({"שם מלא": "כהנא", "רחוב": "השופטים"}),
            Family({"שם מלא": "נתאי", "רחוב": "סלינג'ר"})]
        
        test_cases = [
            ("None", None, 3),
            ("Empty String", "", 3),
            ("Two Matches", "ם", 2),
            ("One Match", "סלי", 1),
            ("Second Letter", "ב", 1),
            ("Not Found", "ברוזוביץ", 0)
        ]

        for title, query, expected_len in test_cases:
            with self.subTest(title=title):
                write_families_file(families=families)
                families_file = load_families_file()
                search_result = search_families(families_file, query, 'street')
                self.assertEqual(expected_len, len(search_result))

    def test_search_by_phone(self):
        families = [Family({"שם מלא": "פרינץ", "מס' פלאפון": "333-4444444", "מס' בית": "000-1111111"}),
            Family({"שם מלא": "כהנא", "מס' פלאפון": "444-5555555", "מס' בית": "111-2222222"}),
            Family({"שם מלא": "נתאי", "מס' פלאפון": "555-6666666", "מס' בית": "222-3333333"})]
        
        test_cases = [
            ("None", None, 3),
            ("Empty String",        "", 3),
            ("Home: One Match",     "0", 1),
            ("Home: Two Matches",   "1", 2),
            ("Mobile: One Match",   "6", 1),
            ("Mobile: Two Matches", "5", 2),
            ("Both: Two Matches",   "3", 2),
            ("Not Found",           "7", 0)
        ]

        for title, query, expected_len in test_cases:
            with self.subTest(title=title):
                write_families_file(families=families)
                families_file = load_families_file()
                search_result = search_families(families_file, query, 'phone')
                self.assertEqual(expected_len, len(search_result))

    def test_search_by_driver(self):
        families = [Family({"שם מלא": "פרינץ", "נהג": "ארז משה"}),
            Family({"שם מלא": "כהנא", "נהג": "אלי לולו"}),
            Family({"שם מלא": "נתאי", "נהג": "יוסי זהר"})]
        
        test_cases = [
            ("None", None, 3),
            ("Empty String", "", 3),
            ("Two Matches", "א", 2),
            ("One Match", "מ", 1),
            ("Not First Letter", "ר", 2),
            ("Not Found", "חיים", 0)
        ]

        for title, query, expected_len in test_cases:
            with self.subTest(title=title):
                write_families_file(families=families)
                families_file = load_families_file()
                search_result = search_families(families_file, query, 'driver')
                self.assertEqual(expected_len, len(search_result))

    def test_search_no_value_cell(self):
        # Should not return this cell for all queries
        families = [Family({"שם מלא": "פרינץ", "רחוב": None})]
        
        test_cases = [
            ("None", None, 0),
            ("Empty String", "", 0),
            ("Not Found", "פ", 0)
        ]

        for title, query, expected_len in test_cases:
            with self.subTest(title=title):
                write_families_file(families=families)
                families_file = load_families_file()
                search_result = search_families(families_file, query, 'street')
                self.assertEqual(expected_len, len(search_result))

    def test_search_hyphen(self):
        families = [Family({"שם מלא": "פרינץ", "רחוב": "שפרינצק", "מס' בית": "000-1111111"})]
        
        test_cases = [
            ("Name Without Hyphen", "פר",   "name", 1),
            ("Name With Hyphen", "פ-ר",     "name", 0),
            ("Street Without Hyphen", "צק", "street", 1),
            ("Street With Hyphen", "צ-ק",   "street", 0),
            ("Phone Without Hyphen", "01",  "phone", 1),
            ("Phone With Hyphen", "0-1",    "phone", 1),
        ]

        for title, query, search_by, expected_len in test_cases:
            with self.subTest(title=title):
                write_families_file(families=families)
                families_file = load_families_file()
                search_result = search_families(families_file, query, search_by)
                self.assertEqual(expected_len, len(search_result))

class TestDataManagement(unittest.TestCase):
    def setUpClass():
        copy(families_filename, 'temp.xlsx')
    
    def tearDownClass():
        copy('temp.xlsx', families_filename)
        remove('temp.xlsx')

    def test_get_families_num(self):
        expected_count = 10
        families = [Family({"שם מלא": f"משפחה מס' {i}"}) for i in range(expected_count)]
        write_families_file(families)
        families_file = load_families_file()
        self.assertEqual(expected_count, get_count(families_file))

    def test_add_family(self):
        exist_families = [Family({"שם מלא": "שלום פרינץ"})]

        test_cases = [
            ("New Name",                {"שם מלא": "נתאי פרינץ"},                      add_results["FAMILY_ADDED"]),
            ("Name Exists",             {"שם מלא": "שלום פרינץ"},                      add_results["FAMILY_EXISTS"]),
            ("Name Partially Exists",   {"שם מלא": "שלום"},                             add_results["FAMILY_ADDED"]),
            ("Missing Name",            {},                                              add_results["MISSING_FULL_NAME"]),
            ("Phone Not Digits",        {"שם מלא": "א", "מס' בית": "שלוםפרינץ"},       add_results["PHONE_NOT_DIGITS"]),
            ("Phone Too Short",         {"שם מלא": "א", "מס' פלאפון": "05253816"},      add_results["PHONE_WRONG_LEN"]),
            ("Phone Too Long",          {"שם מלא": "א", "מס' בית": "05253816480"},      add_results["PHONE_WRONG_LEN"]),
            ("Phone OK",                {"שם מלא": "א", "מס' פלאפון": "0525381648"},    add_results["FAMILY_ADDED"]),
            ("Phone With Hyphen",       {"שם מלא": "א", "מס' בית": "04-5381648"},       add_results["FAMILY_ADDED"]),
            ("Phone With Hyphen",       {"שם מלא": "א", "מס' פלאפון": "052-5381648"},   add_results["FAMILY_ADDED"]),
        ]

        for title, family, expected_result in test_cases:
            with self.subTest(title=title):
                write_families_file(families=exist_families)
                families_file = load_families_file()
                result = add_family(families_file, family)
                self.assertEqual(expected_result, result)

    def test_add_many_families(self):
        family = {"שם מלא": "שלום פרינץ"}
        exist_families = [Family(family)]
        exists_error = add_many_error(add_results["FAMILY_EXISTS"], family["שם מלא"])

        test_cases = [
            ("No Families",             [],                                               add_many_results["FAMILIES_ADDED"]),
            ("New Name",                [{"שם מלא": "נתאי פרינץ"}],                      add_many_results["FAMILIES_ADDED"]),
            ("Name Exists",             [{"שם מלא": "שלום פרינץ"}],                      exists_error),
            ("Second Family Exists",    [{"שם מלא": "שלום"}, {"שם מלא": "שלום פרינץ"}], exists_error),
        ]

        for title, families, expected_result in test_cases:
            with self.subTest(title=title):
                write_families_file(families=exist_families)
                families_file = load_families_file()
                result = add_families(families_file, families)
                self.assertEqual(expected_result, result)

    def test_add_families_before_error(self):
        family = {"שם מלא": "שלום פרינץ"}
        exist_families = [Family(family)]
        write_families_file(families=exist_families)
        
        families_file = load_families_file()
        families = [{"שם מלא": "דוד חיים"}, {"שם מלא": "משפוחה"}, {"שם מלא": "שלום פרינץ"}]
        result = add_families(families_file, families)
        
        exists_error = add_many_error(add_results["FAMILY_EXISTS"], family["שם מלא"])
        self.assertEqual(exists_error, result)

        first_search = search_families(families_file, "דוד חיים", 'name')
        self.assertEqual(1, len(first_search))
        second = search_families(families_file, "משפוחה", 'name')
        self.assertEqual(1, len(second))
